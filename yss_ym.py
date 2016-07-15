import os
import shutil
import xlrd
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter

source_list = []
source_path = r'C:\PythonProject\pythonexcel'
for dir_path, subdir_path, files in os.walk(source_path):
    # print (dir_path, subdir_path, files)
    files.remove(r'yss_ym.xlsx')
    files.remove(r'保本理财利润表.et')
    files.remove(r'保本理财利润表.xlsx')
    files.remove(r'yss_ym.py')
    # files.remove(r'history')
    for file in files:
        file_path = os.path.join(dir_path, file)
        source_list.append(file_path)

wb_comb = load_workbook(r'yss_ym.xlsx', guess_types=False)
ws_comb = wb_comb.active

for i in range(0, len(source_list)):
    ws_comb.cell(row=5, column=i + 4).value = files[i]
    ws_comb.cell(row=5, column=i + 4 + len(source_list)).value = files[i]
    # 第五行列明原始表名
    # 遍历赋值
    for j in range(6, 35):
        ws_comb.cell(row=j, column=i + 4).value = xlrd.open_workbook(source_list[i]).sheets()[0].cell_value(j - 1, 1)
        ws_comb.cell(row=j, column=i + 4 + len(source_list)).value =\
            xlrd.open_workbook(source_list[i]).sheets()[0].cell_value(
            j-1, 2)
        # 获取列数对应的字母
        end_col_of_thisterm = get_column_letter(len(source_list)+3)
        start_col_of_lastterm = get_column_letter(len(source_list)+4)
        end_col_of_lastterm = get_column_letter(len(source_list)*2+3)
        # 添加汇总合计公式在B、C列
        ws_comb.cell(row=j, column=2).value = '=ROUND(SUM(D%s:%s%s)/100000000.00,2)' % (j, end_col_of_thisterm, j)
        ws_comb.cell(row=j, column=3).value = '=ROUND(SUM(%s%s:%s%s)/100000000.00,2)' % (start_col_of_lastterm, j,
                                                                                         end_col_of_lastterm, j)
wb_comb.save(r'yss_ym.xlsx')

for abssource in source_list:

    shutil.copy(abssource, r'C:\PythonProject\history')
    os.remove(abssource)
